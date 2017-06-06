#!/usr/bin/env python
"""
Auxiliary program for reading, analyzing and writing data of poses and time.

This module allows:
-Read data of poses with their respective time, from a spreadsheet.
-Read data directly from an array of float32 Nx4.
-Analyze data by calculating differential values of position, time and lenght
 displaced and average speed of UGV
-Save data in spreadsheet and textfile.
-Save final data in master sheet.
"""
# Standard libraries
import glob
import math
import numpy as np
import os
from scipy import stats
import sys
import time
#Excel read/write library
import openpyxl

def format_spreadsheet(cell):
    """
    It allows different types of format in spreadsheet cells.

    :param cell: type of format desired.
    """
    #Types of text alignment.
    if cell == 'center_al':
        format_cell = openpyxl.styles.Alignment(horizontal='center',
                                                vertical='center')
    elif cell == 'right_al':
        format_cell = openpyxl.styles.Alignment(horizontal='right',
                                                vertical='center')
    #Types of font.
    elif cell == 'title_ft':
        format_cell = openpyxl.styles.Font(color='FFFFFFFF', size=20, bold=True)
    elif cell == 'white_ft':
    #Types of border.
        format_cell = openpyxl.styles.Font(color='FFFFFFFF', bold=True)
    elif cell == 'thick_bd':
        format_cell = openpyxl.styles.Border(
                                top=openpyxl.styles.Side(border_style='thick',
                                                         color='FF4143CA'),
                                bottom=openpyxl.styles.Side(border_style='thick',
                                                            color='FF4143CA'))
    elif cell == 'thin_bd':
        format_cell = openpyxl.styles.Border(
                                top=openpyxl.styles.Side(border_style='thin',
                                                         color='FF83C6D6'),
                                bottom=openpyxl.styles.Side(border_style='thin',
                                                            color='FF83C6D6'))
    #Types of fill.
    elif cell == 'blue_fill':
        format_cell = openpyxl.styles.PatternFill(fill_type='solid',
                                                  start_color='FF2F79E6',
                                                  end_color='FF2F79E6')
    elif cell == 'skyblue_fill':
        format_cell = openpyxl.styles.PatternFill(fill_type='solid',
                                                  start_color='FFDBF4FA',
                                                  end_color='FFDBF4FA')
    elif cell == 'white_fill':
        format_cell = openpyxl.styles.PatternFill(fill_type='solid',
                                                  start_color='FFFFFFFF',
                                                  end_color='FFFFFFFF')
    return format_cell

def read_data(filename_spreadsheet="datatemp/31_05_2017_61-L160-R210.xlsx", analyze=True):
    """
    It allows to read poses and time of spreadsheet to analyze or save them.

    :param filename_spreadsheet: name of spreadsheet that contain the data to
    be read.
    """
    ###29_05_2017_2-L160-R160.xlsx
    try:
        wb = openpyxl.load_workbook(filename_spreadsheet)
    except IOError:
        wb = openpyxl.Workbook()
    ws = wb.active
    #Initialization of matrices for data.
    data = np.array([0, 0, 0, 0]).astype(np.float64)
    last_data = np.array([0, 0, 0, 0]).astype(np.float64)
    new_data = np.array([0, 0, 0, 0]).astype(np.float64)
    #The first row is the header, and the second is a set of zeros (already
    #initialized in the matrix). Begins to read row 3.
    row = 7
    #Number of columns in the matrix.
    cols = data.shape[0]
    #Loop for reading data.
    current_row_data = True
    while current_row_data:
        element = ws.cell(column=1, row=row).value
        if element == 'Sum differential data:':
            current_row_data = False
        else:
            new = False
            #import pdb; pdb.set_trace()
            for y in range (0, cols):
                element = ws.cell(column=y+1, row=row).value
                new_data[y] = element
                if y > 0 :
                    if new_data[y] != last_data [y]:
                        new = True
            if new == True:
                last_data = np.copy(new_data)
                data = np.vstack([data, new_data])
            row +=1
    formatted_data = np.round(data, 2)
    #Call to save and analyze data.
    save_data(filename_spreadsheet, formatted_data, analyze=analyze)
    return formatted_data
def save_data(filename_spreadsheet, data, analyze=False):
    """
    Receives poses and time of matrix to analyze and/or save them.

    :param data: Matrix of floats64 with data.
    :param analyze: Boolean that allows analyze or not the data.
    """
    #Get the SP values from the user.
    time.sleep(0.2)
    #Delete first row data (row of zeros).
    data = data[1:data.shape[0], :]
    #First sample, time zero.
    data[0:data.shape[0], 0] = data[0:data.shape[0], 0] - data[0, 0]
    #TODO Try, except correct value.
    sp_left = input("Introduce value of sp_left between 0 and 255\n")
    sp_right = input("Introduce value of sp_right between 0 and 255\n")
    #Header construction and data analysis if the latter is required.
    if analyze:
        #Call for data analysis function.
        data, save_master, avg_speed, avg_ang_speed = analyze_data(data)
        header_text = np.array(['Time', 'Pos x', 'Pos y', 'Angle', 'Diff Time',
                            'Diff Posx', 'Diff Posy', 'Diff Angl', 'Diff Leng',
                            'Rel Speed', 'Rel AnSpd'])
    else:
        header_text = np.array(['Time', 'Pos x', 'Pos y', 'Angle'])
        save_master = True

    full_data = np.vstack([header_text, data])
    # Name of the output file for the poses historic values.
    exist_file = glob.glob("datatemp/*.xlsx")
    exist_file.sort()
    index = len (exist_file)
    datestamp = "RW{}".format(time.strftime("%d_%m_%Y"))
    filename = "{}_{}-L{}-R{}".format(datestamp, (index+1), sp_left, sp_right)
    datestamp = '{}_{}'.format(datestamp, (index+1))
    name_txt = "datatemp/{}.txt".format(filename)
    name_mastertxt = "datatemp/masterfile2.txt"
    #Header for numpy savetxt.
    header_numpy = ''
    cols = header_text.shape[0]
    for x in range (0, cols):
        element = header_text[x]
        element = '%9s' % (element)
        header_numpy = '{}{}\t'.format(header_numpy, element)
    #Call to save data in textfile.
    np.savetxt(name_txt, data, delimiter='\t', fmt='%9.2f',
               header=header_numpy, comments='')
    #Experiment conditions.
    exp_conditions = (" -Use camera 3\n -Position initial experiment forward: "
                      "right rear wheel profile (-1400, -600), rear axis UGV in "
                      "axis y, in -1800 x\n -Time: 3 seconds")
    #Call to save data in spreadsheet.
    name_to_use = data2spreadsheet(header_text, full_data, filename,
                                   exp_conditions, save_master)
    #Save data to masterfile.
    if save_master:
        #Call to save data in spreadsheet masterfile.
        data_master = np.array([datestamp, sp_left, sp_right, name_to_use])
        save2master_xlsx(data_master)
        #Call to save data in text masterfile.
        data_master_txt = np.array([datestamp, sp_left, sp_right, avg_speed,
                                    avg_ang_speed])
        save2master_txt(data_master_txt)

def analyze_data(data):
    """
    Receives poses and time of matrix to analyze.

    Different time and position values are calculated between two data points.
    From these, the displaced length and the average speed of UGV are calculated.

    :param data: Matrix of floats64 with data.
    """
    rows, cols = data.shape
#     #Data erase with UGV stopped.
#     #Initial repeated data calculation.
#     pos_x_upper = data[0:20, 1]
#     mode_pos_x_upper = stats.mode(pos_x_upper)
#     pos_y_upper = data[0:20, 2]
#     mode_pos_y_upper = stats.mode(pos_y_upper)
#     #Final repeated data calculation.
#     pos_x_lower = data[(rows-20):rows, 1]
#     mode_pos_x_lower = stats.mode(pos_x_lower)
#     pos_y_lower = data[(rows-20):rows, 2]
#     mode_pos_y_lower = stats.mode(pos_y_lower)
#     #Determination of rows UGV data in motion.
# #conditions = np.any(data!=moda, axis=1)
# #indexes = np.where(conditions)[0]
# #filtered_data = data[indexes[0]-1:indexes[1]+2, :]
    row_upper = 0
    row_lower = rows
    # for x in range(0, rows):
    #     if data[x,1] == mode_pos_x_upper[0] and data[x,2] == mode_pos_y_upper[0]:
    #         row_upper = x
    #     if data[x,1] == mode_pos_x_lower[0] and data[x,2] == mode_pos_y_lower[0]:
    #         row_lower = x + 1
    #         break
    # #UGV data in motion.
    clipped_data = data[row_upper:row_lower, :]
    rows, cols = clipped_data.shape
    #First sample, time zero.
    clipped_data[:, 0] -= clipped_data[0, 0]
    #Differential data matrix: current data minus previous data.
    diff_data = np.zeros_like(clipped_data)
    #Vector differential angle speed.
    diff_angle_speed = np.zeros(rows)
    diff_data[1:rows, :] = clipped_data[1:rows, :] - clipped_data[0:(rows-1), :]
    #Vector differential length displaced.
    diff_length = np.sqrt(diff_data[:,1] ** 2 + diff_data[:,2] ** 2)
    speed_angles = np.arctan2(diff_data[:,2], diff_data[:,1])
    # The direction is negative when the speed angle and vehicle angle difference
    # is bigger than pi/2 (90 degrees).
    sign_spd = np.ones([rows, 1])
    sign_spd[np.abs(clipped_data[:,3]-speed_angles) > (np.pi/2)] *= -1
    diff_speed = sign_spd[1:,0] * 1000 * diff_length[1:]/diff_data[1:,0]
    diff_speed = np.hstack([0, diff_speed])
#    for x in range(1, rows):
#        cos_alpha[x] = diff_data[x,1] / diff_length[x]
#        cos_thetha[x] = math.cos(data[x,3])
#        sign_spd[x] = math.sign(cos_thetha[x]/cos_alpha[x])
#        if diff_data[x,0] != 0:
#            #Speed in millimeters/second.
#            diff_speed[x] = ((sign_spd[x] * diff_length[x]) / diff_data[x,0]) * 1000
    diff_angle_speed[1:rows] = 1000 * diff_data[1:rows, 3]/diff_data[1:rows, 0]
    #Complete data matrix with new data.
#    np.hstack([diff_data[4], diff_length])
    diff_data = np.insert(diff_data, 4, diff_length, axis=1)
    diff_data = np.insert(diff_data, 5, diff_speed, axis=1)
    diff_data = np.insert(diff_data, 6, diff_angle_speed, axis=1)
    clipped_data = np.hstack([clipped_data, diff_data])
    #import pdb; pdb.set_trace()
    #Average speed and average angle speed for masterfile txt.
    rows, cols = clipped_data.shape
    sum_data = diff_data.sum(axis=0)
    length = math.sqrt(((clipped_data[rows-1,1] - clipped_data [0, 1])**2)+((clipped_data[rows-1,2] - clipped_data [0, 2])**2))
    sum_data = diff_data.sum(axis=0)
    avg_speed = np.round((1000*length/sum_data[0]), 2)
    avg_ang_speed=np.round(((1000*(clipped_data[rows-1,3] - clipped_data [0, 3]))/sum_data[0]), 2)
    #If you want to save to master file boolean True.
    save_master = True
    formatted_data = np.round(clipped_data, 2)
    return formatted_data, save_master, avg_speed, avg_ang_speed

def data2spreadsheet(header, data, filename, exp_conditions, save_master):
    """
    Receives poses and time, and saves them in a spreadsheet.

    :param header: contains header to save in spreadsheet.
    :param data: contains data to save in spreadsheet.
    :param filename: name of spreadsheet where the data will be saved.
    :param exp_conditions: contains string with experiment description.
    :param save_master: boolean with True for save data in masterfile.
    """
    name_spreadsheet = "datatemp/{}.xlsx".format(filename)
    try:
        wb = openpyxl.load_workbook(name_spreadsheet)
    except:
        wb = openpyxl.Workbook()
    ws = wb.active
    #Spreadsheet title.
    ws.merge_cells('A1:K2')
    ws.cell('A1').value = filename
    ws.cell('A1').alignment = format_spreadsheet('center_al')
    ws.cell('A1').font = format_spreadsheet('title_ft')
    ws.cell('A1').fill = format_spreadsheet('blue_fill')
    #Experiment conditions.
    ws.merge_cells('A3:K5')
    ws.cell('A3').value = exp_conditions
    #Freeze header
    my_cell = ws['B7']
    ws.freeze_panes = my_cell
    #Write in spreadsheet the headboard.
    rows = data.shape[0]
    cols = data.shape[1]
    for y in range (0, cols):
        ws.cell(column=y+1, row=6, value=header[y])
        ws.cell(column=y+1, row=6).alignment = format_spreadsheet('right_al')
        ws.cell(column=y+1, row=6).font = format_spreadsheet('white_ft')
        ws.cell(column=y+1, row=6).fill = format_spreadsheet('blue_fill')
        ws.cell(column=y+1, row=6).border = format_spreadsheet('thick_bd')
    #Write in spreadsheet the data.
    for x in range(1, rows):
        for y in range(0, cols):
            element = float(data[x,y])
            ws.cell(column=y+1, row=x+6, value=element).number_format = '0.00'
            if x % 2 != 0:
                ws.cell(column=y+1, row=x+6).fill = format_spreadsheet(
                                                                 'skyblue_fill')
            else:
                ws.cell(column=y+1, row=x+6).fill = format_spreadsheet(
                                                                   'white_fill')
            my_cell = ws.cell(column=y+1, row=x+6)
            ws.column_dimensions[my_cell.column].width = 10
    #Write in spreadsheet the name of statistics.
    for x in range(rows+6, rows+12):
        ws.merge_cells(start_row=x,start_column=1,end_row=x,end_column=4)
        ws.cell(column=1, row=x).alignment = format_spreadsheet('right_al')
        ws.cell(column=1, row=x).fill = format_spreadsheet('blue_fill')
        ws.cell(column=1, row=x).font = format_spreadsheet('white_ft')
    ws.cell(column=1, row=rows+6, value='Sum differential data:')
    ws.cell(column=1, row=rows+7, value='Mean of differential data:')
    ws.cell(column=1, row=rows+8, value='Variance differential data:')
    ws.cell(column=1, row=rows+9, value='Std deviation differential data:')
    ws.cell(column=8, row=rows+10, value='Linear Relative Speed:')
    ws.cell(column=8, row=rows+11, value='Angular Relative Speed:')
    ws.merge_cells(start_row=rows+10,start_column=8,end_row=rows+10,end_column=10)
    ws.merge_cells(start_row=rows+11,start_column=8,end_row=rows+11,end_column=10)
    ##Write and calculate in spreadsheet the statistics.
    for y in range(5, cols+1):
        letter_range = openpyxl.utils.get_column_letter(y)
        start_range = '{}{}'.format(letter_range, 7)
        end_range = '{}{}'.format(letter_range, rows+5)
        interval = '{}:{}'.format(start_range,end_range)
        ws.cell(column=y, row=rows+6, value= '=SUM({})\n'.format(interval))
        ws.cell(column=y, row=rows+7, value= '=AVERAGE({})\n'.format(interval))
        ws.cell(column=y, row=rows+8, value= '=VAR({})\n'.format(interval))
        ws.cell(column=y, row=rows+9, value= '=STDEV({})\n'.format(interval))
        for x in range(rows+6, rows+12):
            ws.cell(column=y, row=x).number_format = '0.00'
            ws.cell(column=y, row=x).font = format_spreadsheet('white_ft')
            ws.cell(column=y, row=x).fill = format_spreadsheet('blue_fill')
            ws.cell(column=y, row=x).alignment = format_spreadsheet('right_al')
    ws.cell(column=11, row=rows+10, value= '=1000*SQRT(((B{rows}-B7)^2)+'
            '(((C{rows}-C7)^2)))/E{rows2}\n'.format(rows=rows+5, rows2=rows+6))
    ws.cell(column=11, row=rows+11, value= '=1000*(D{rows}-D7)/'
                                'E{rows2}\n'.format(rows=rows+5, rows2=rows+6))
    wb.save(name_spreadsheet)

    return name_spreadsheet

def save2master_xlsx(data_master):
    """
    Average speed, setpoint left and right wheels are saved in the same spreadsheet.

    :param data_master: array that contains the average lineal speed, the
    average angular speed, the set point left, the set point right, and the
    name of datafile.
    """
    #Data search to save in masterspreadsheet.
    folder = '\'file:///home/joselamas/repository_ugv/UviSpace/uvispace/uvisensor/'
    name_sheet = '\'#$Sheet.'
    try:
        wb = openpyxl.load_workbook(data_master[3])
    except:
        wb = openpyxl.Workbook()
    ws = wb.active
    #Next empty row search.
    row = 7
    written_row = True
    while written_row:
        element = ws.cell(column=1, row=row).value
        if element == None:
            written_row = False
        else:
            row +=1
    avg_speed = folder + data_master[3] + name_sheet + 'K' + '{}'.format(row)
    row = row + 1
    avg_ang_speed = folder + data_master[3] + name_sheet + 'K' + '{}'.format(row)
    #Save data in masterspreadsheet.
    try:
        wb = openpyxl.load_workbook("datatemp/masterfile2.xlsx")
    except:
        wb = openpyxl.Workbook()
    ws = wb.active
    #Freeze header
    my_cell = ws['A2']
    ws.freeze_panes = my_cell
    #Next empty row search.
    row = 1
    written_row = True
    while written_row:
        element = ws.cell(column=1, row=row).value
        if element == None:
            written_row = False
        else:
            row +=1
    #Save data in empty row.
    ws.cell(column=1, row=row, value=data_master[0])
    ws.cell(column=2, row=row, value=data_master[1])
    ws.cell(column=3, row=row, value=data_master[2])
    ws.cell(column=4, row=row, value= '=INDIRECT(F{})\n'.format(row)).number_format = '0.00'
    ws.cell(column=5, row=row, value= '=INDIRECT(G{})\n'.format(row)).number_format = '0.00'
    ws.cell(column=6, row=row, value=avg_speed)
    ws.cell(column=7, row=row, value=avg_ang_speed)
    ws.cell(column=8, row=row, value="_")
    #Format data.
    for y in range (1, 8):
        my_cell = ws.cell(column=y, row=row)
        if y == 1:
            ws.column_dimensions[my_cell.column].width = 18
        else:
            ws.column_dimensions[my_cell.column].width = 10
        if row % 2 != 0:
            ws.cell(column=y, row=row).fill = format_spreadsheet('skyblue_fill')
        else:
            ws.cell(column=y, row=row).fill = format_spreadsheet('white_fill')
        if y < 6:
            ws.cell(column=y, row=row).alignment = format_spreadsheet('right_al')
    wb.save("datatemp/masterfile3.xlsx")

def save2master_txt(data_master):
    """
    Average speed, setpoint left and right wheels are saved in the same textfile.

    :param data_master: array that contains the average lineal speed, the
    average angular speed, the set point left, the set point right, and the
    name of datafile.
    """
    #TODO improve format
    cols = data_master.shape[0]
    for y in range (0, cols):
        value = data_master[y]
        if y == 0:
            #TODO use '.format' string construction style
            text = '%9s' % (value)
        else:
            value = float(value)
            value = '%9.2f' % (value)
            text = '{}\t\t{}'.format(text, value)
    text = '{}\n'.format(text)
    with open("datatemp/masterfile3.txt", 'a') as outfile:
        outfile.write(text)


def main():
    o_exist_file = glob.glob("./datatemp/*.xlsx")
    o_exist_file.sort()
    o_index = len (o_exist_file)
    names = [os.path.basename(x) for x in o_exist_file]
    for y in range (0, o_index-2):
        #import pdb; pdb.set_trace()
        print names[y]
        read_data(filename_spreadsheet='datatemp/' + names[y], analyze=True)

        import re

test_string = "29_05_2017_16-L160-R200.xlsx"
# Use as the beginning of the pattern the L letter, and the '-' as the end.
pattern = '-L.*-'
match = re.search(pattern, test_string)
left_speed = match[2:-1]
# Use as the beginning of the pattern the R letter, and the '-' as the end.
pattern = '-R.\.'
match = re.search(pattern, test_string)
right_speed = match[2:-1]

if __name__ == '__main__':
    main()
